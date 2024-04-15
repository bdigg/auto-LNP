import pump
import sys
import os
import time
import matplotlib.pyplot as plt
import numpy as np
import threading
import expel
import pandas as pd
import math
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime


def flush(active_channels, pressure, last_t):
    print("Flushing System")
    for i, channel in enumerate(active_channels):
        start_t = time.time() 
        error = pump.set_pressure(channel,pressure)
        print("Channel",channel,"Pressure",pressure)
    while True:
        if (time.time() - start_t) > last_t:
            break
    return error

def stability_test(active_channels, pressures): #update to give live plot and value for stability
    print("Stability Test")
    for pressure in pressures:
        for channel in active_channels:
            error = pump.set_pressure(channel,pressure)
            print("Channel",channel,"Pressure",pressure)
        print("Wait 10 seconds")
        time.sleep(10)
        start_t,last_t = time.time(),time.time()
        exp_t = 5
        t = 0
        prange = [[],[],[],[]]
        while True:
            for i, channel in enumerate(active_channels):
                flowrate = pump.get_sensor_data(channel)
                print(channel,pump.get_sensor_data(channel))
            print(t,"seconds")
            if (time.time() - start_t) > exp_t:
                break
            # Wait until desired period time
            sleep_t = 1 - (time.time() - last_t)
            if sleep_t > 0:
                time.sleep( sleep_t )
            last_t = time.time() # And update the last time  
            t += 1   
    return error      

def stop():
    pump.set_pressure(1,0)
    pump.set_pressure(2,0)
    pump.set_pressure(3,0)
    pump.set_pressure(4,0)

def expulsiontime(tubingdim,set_FR,active_channels):
    diameter = math.pi()*(tubingdim[0]/2)**2
    tubet = (diameter*(tubingdim[2]*10))/np.sum(set_FR)
    return tubet


def savetoexcel(exp_name,status,expparams,exp_FRs,wpindex,volume,fr_perc_error,repeat,eq_t):
    data = {
        'ExpName': [exp_name],
        'State': [status],
        'RepeatNum': [repeat],
        'Time': [datetime.now().strftime('%H:%M:%S')],
        'Date': [(datetime.today().strftime('%Y-%m-%d'))],
        'WPIndex': [str(wpindex)],
        'FRR': [expparams[1]],  
        'TotalFR': [expparams[0]],  
        'Volume': [volume],
        'Eq Time': [eq_t],
        'Buf-Name': [expparams[5]],
        'Buf-FR': [exp_FRs[0]],
        'Buf-FRer': [fr_perc_error[0]],
        'Lp1-Name': [expparams[6]],
        'Lp1-Comp': [expparams[2]],
        'Lp1-FR': [exp_FRs[1]],
        'Lp1-FRer': [fr_perc_error[1]],
        'Lp2-Name': [expparams[7]],
        'Lp2-Comp': [expparams[3]],
        'Lp2-FR': [exp_FRs[2]],
        'Lp2-FRer': [fr_perc_error[2]],
        'Lp3-Name': [expparams[8]],
        'Lp3-Comp': [expparams[4]],
        'Lp3-FR': [exp_FRs[3]],
        'Lp3-FRer': [fr_perc_error[3]],        
    }

    # Convert data to DataFrame
    df = pd.DataFrame(data)

    # Excel file name
    excel_file = 'ExperimentLog.xlsx'

    try:
        # Load the existing workbook
        book = load_workbook(excel_file)
        
        # Select the active sheet
        sheet = book.active
        
        # Find the next available row
        next_row = sheet.max_row + 1
        
        # Append DataFrame to the existing sheet
        for r in dataframe_to_rows(df, index=False, header=False):
            sheet.append(r)
        
        # Save the workbook
        book.save(excel_file)
        
        print("Data appended to Excel successfully.")
        
    except FileNotFoundError:
        # If the file doesn't exist, create a new Excel file and write the DataFrame
        df.to_excel(excel_file, index=False)
    print("Data written to a new Excel file successfully.")

#-----------------------------  

def plotupdate(ax1,ax2,flow_data,active_channels,col_list,set_FR,p_range,n,k,single):
    ax1.clear()
    ax2.clear()
    chancolours = ["lightseagreen","darkorange","darkred","blueviolet"]
    for i, chan in enumerate(active_channels):
        ax1.plot(flow_data[0][n:],flow_data[1][i][n:],label="Channel %d" % chan, color=chancolours[i])
        ax2.plot(flow_data[0][n:],flow_data[3][i][n:],label="Channel %d" % chan,color=chancolours[i])
        if single == True:
            ax1.axhline(y=set_FR[i],color=chancolours[i],linewidth=0.5, alpha=0.5)
    if col_list > 0:
        ax1.axvline(x=col_list,color='k',linestyle=':')
        ax2.axvline(x=col_list,color='k',linestyle=':')                    
    ax1.set_ylim(0, np.amax(set_FR)*1.5)
    ax2.set_ylim(0,p_range[1])
    ax1.legend()
    ax2.legend()
    ax1.set_ylabel('flow [uL/min]')
    ax2.set_ylabel('pressure [mbar]')
    plt.xlabel('Time (s)')
    plt.pause(0.0001)
    plt.show(block=False)

#------------------
                  
def main_PI(expname,exp_params,autocollect,active_channels,period,K_p,K_i,exp_FRs,volume,p_range,p_incr,eqb_max,eqb_min,wpdim,wpcurrent,tubingdim,standard_repeats,ser):
    
    flow_data,n = [[],[[],[],[],[]],[[],[],[],[]],[[],[],[],[]]],0 #time,flow,realp,setp

    #create folder for saving images
    path = './Flowplots/' + expname
    if not os.path.exists(path):
        os.makedirs(path)
        print(f"Directory '{path}' created successfully.")
    else:
        print(f"Directory '{path}' already exists.")

    #Global PID variables
    pr_control, I = [0,0,0,0],[0,0,0,0]
    wpprev = [1,1]

    #Setup live flow figure
    fig = plt.figure()
    ax1 = fig.add_subplot(2,1,1)
    ax2 = fig.add_subplot(2,1,2)
    fr = [0,0,0,0]

    init_start_t = time.time()
    timelast = time.time()

    moved = True


    #Iteration through each of the flow rates
    for j, set_FR in enumerate(exp_FRs): 

        repeats = 0
        col_list = 0
        
        exp_name = str(exp_params[1]) + "-" + expname + str(exp_params[2]) + str(exp_params[3]) + str(exp_params[4] )
        print(exp_name)
        #caclulate experiment time based off the target volume
        exp_t = (volume*1.2/np.sum(set_FR[0:(len(active_channels)-1)]))*60

        #tubevol = ((math.pi)*(tubingdim[0]/2)**2)*(tubingdim[1]*10)

        while repeats < standard_repeats: #Gives the option to repeat a given flowrate condition
            
            print("Beginning experiment: ", set_FR)
            
            fr_perc_error, max_fr_error = [[0],[0],[0],[0]],[]
            collection,consistent_fr = False,False
            
            wpprev = [wpcurrent[0],wpcurrent[1]]
            if wpcurrent[1] == wpdim[1]:
                wpcurrent[1] = 1
                wpcurrent[0] = wpcurrent[0] + 1
            else:
                wpcurrent[1] = wpcurrent[1] + 1

            #Read inital time
            start_t = time.time() 
            last_t = start_t
            active_t = eqb_max           

            while True:
                #-----Main PID Loop-----
                flow_data[0].append((time.time() - init_start_t))

                #Repeat for all active channels
                for i, channel  in enumerate(active_channels):
                    #Flow rate reading
                    fr[i] = pump.get_sensor_data(channel)[0]
                    if channel > 1: #If a lipid channel (in ethanol) make adjustment
                        fr[i] = (fr[i] - 80.09)*4.05
                    flow_data[1][i].append(fr[i])

                    #Calculate error
                    fr_error = fr[i]-set_FR[i]
                    fr_perc_error[i].append(abs(fr_error/set_FR[i]))

                    #Pressure reading
                    pr = pump.get_pressure_data(channel)[0]  
                    flow_data[2][i].append(pr)

                    #Correction calculations
                    I[i] = I[i] + K_i*fr_error*(period)
                    adjustment =  (fr_error*abs(fr_error))*K_p[i] + I[i]

                    #Set limits
                    adjustment = np.clip(adjustment, p_incr[0], p_incr[1])
                    pr_control[i] = pr_control[i] - adjustment 
                    pr_control[i] = np.clip(pr_control[i], p_range[0], p_range[1])
                    flow_data[3][i].append(pr_control[i])

                    #Set new pressure
                    pump.set_pressure(channel,pr_control[i]) 
                 
                max_fr_error.append(np.max([errors[-1] for errors in fr_perc_error]))

                if (time.time() - start_t) > eqb_min and collection == False:
                    if moved == False:
                         #Move to coord for collection
                        #Move to new coordinate 
                        threading.Thread(target=expel.nextwell,args=(ser, wpprev, wpcurrent)).start()
                        print("Current plate position: ",wpcurrent)
                        moved = True

                    if np.max(max_fr_error) < 0.05:
                        print("FR condition reached \n\n\n")
                        #Move to well position
                        if autocollect == True:
                            expel.flowswitch(ser,1)
                        collection = True
                        col_list = time.time() - init_start_t
                        eq_t = time.time() - start_t
                        active_t = exp_t + time.time() - start_t
                        fr_perc_error, max_fr_error = [[0],[0],[0],[0]],[]
                    max_fr_error = max_fr_error[1:] #Remove the first value 
    
                elif np.max(max_fr_error) > 0.1 and collection == True:
                    consistent_fr = False
                    print("Flow rate condition fail")
                        
                if (time.time() - start_t) > active_t:
                    expel.flowswitch(ser,0)
                    break


                # Wait until desired period time
                sleep_t = period - (time.time() - last_t)
                if sleep_t > 0:
                    time.sleep( sleep_t )
                last_t = time.time() # And update the last time 
                
                #Updating Figure
                plotupdate(ax1,ax2,flow_data,active_channels,col_list,set_FR,p_range,n,-1,True)    

            if collection == False:
                status = "Failed to Eq"
                eq_t = 0
                print("Equilibration Failed - FR percentage error", np.max(max_fr_error))
            else:
                if consistent_fr == True: 
                    status = "Complete"
                    print("Collection Successful - FR percentage error", np.max(max_fr_error)) 
                    standard_repeats += 1
                else:
                    status = "Failed"
                    print("Collection Unsuccessful - FR percentage error", np.max(max_fr_error)) 
            
            repeats+= 1

            n = len(flow_data[0])-1
            ch_fr_error = [np.max(fr_perc_error[0]),np.max(fr_perc_error[1]),np.max(fr_perc_error[2]),np.max(fr_perc_error[3])]
            savetoexcel(exp_name,status,exp_params,set_FR,wpcurrent,volume,ch_fr_error,standard_repeats,eq_t)
            plt.savefig(path+exp_name+".png")
            #save wp positions with exp info
            moved = False
            
    stop() 
    plotupdate(ax1,ax2,flow_data,active_channels,col_list,exp_FRs,p_range,0,0,False)    
    plt.show()
    plt.savefig(path+expname+"/.png")

    return()
    
