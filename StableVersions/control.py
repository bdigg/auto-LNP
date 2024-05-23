import pump
import sys
import os
import time
import matplotlib.pyplot as plt
import numpy as np
import threading
import expel
import pandas as pd
import math
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
import random 


def flush(active_channels, pressure, last_t,calibarr):
    print("Flushing System")
    for i, channel in enumerate(active_channels):
        start_t = time.time() 
        error = pump.set_pressure(channel,pressure,calibarr)
        print("Channel",channel,"Pressure",pressure)
    while True:
        if (time.time() - start_t) > last_t:
            break
    return error

def stability_test(active_channels, pressures,calibarr): #update to give live plot and value for stability
    print("Stability Test")
    for pressure in pressures:
        for channel in active_channels:
            error = pump.set_pressure(channel,pressure,calibarr)
            print("Channel",channel,"Pressure",pressure)
        print("Wait 10 seconds")
        time.sleep(10)
        start_t,last_t = time.time(),time.time()
        exp_t = 5
        t = 0
        prange = [[],[],[],[]]
        while True:
            for i, channel in enumerate(active_channels):
                flowrate = pump.get_sensor_data(channel)
                print(channel,pump.get_sensor_data(channel))
            print(t,"seconds")
            if (time.time() - start_t) > exp_t:
                break
            # Wait until desired period time
            sleep_t = 1 - (time.time() - last_t)
            if sleep_t > 0:
                time.sleep( sleep_t )
            last_t = time.time() # And update the last time  
            t += 1   
    return error      

def stop(calibarr):
    pump.set_pressure(1,0,calibarr)
    pump.set_pressure(2,0,calibarr)
    pump.set_pressure(3,0,calibarr)
    pump.set_pressure(4,0,calibarr)

def expulsiontime(tubingdim,set_FR,active_channels):
    diameter = math.pi()*(tubingdim[0]/2)**2
    tubet = (diameter*(tubingdim[2]*10))/np.sum(set_FR)
    return tubet

def saveflowdata(fr_perc_error,flow_data,flowstart,fdataall,path,expname):
    ch_fr_error = [np.max(fr_perc_error[0]),np.max(fr_perc_error[1]),np.max(fr_perc_error[2]),np.max(fr_perc_error[3])]
    mean_fr = [np.mean(flow_data[1][0][flowstart:]),np.mean(flow_data[1][1][flowstart:]),np.mean(flow_data[1][2][flowstart:]),np.mean(flow_data[1][3][flowstart:])]
    stdev_fr = [np.std(flow_data[1][0][flowstart:]),np.std(flow_data[1][1][flowstart:]),np.std(flow_data[1][2][flowstart:]),np.std(flow_data[1][3][flowstart:])]
    fdata = [flow_data[1][0][flowstart:],flow_data[1][1][flowstart:],flow_data[1][2][flowstart:],flow_data[1][3][flowstart:]]
    fdataall.append(fdata)
    DF = pd.DataFrame(fdataall)
    DF.to_csv(path+"/flowdata/"+expname+"-steadyflowonly.csv")
    flowdf = pd.DataFrame(flow_data)
    flowdf.to_csv(path+"/flowdata/"+expname+"-Flowdata.csv")
    return(ch_fr_error,mean_fr,stdev_fr)
    
def savetoexcel(path,expname,exp_name,status,expparams,exp_FRs,wpindex,volume,fr_perc_error,mean_fr,stdev_fr,repeat,eq_t):
    TotalFR = np.sum(exp_FRs)
    FRR = exp_FRs[0]/np.sum(exp_FRs[1:])
    data = {
        'ExpName': [exp_name],
        'State': [status],
        'RepeatNum': [repeat],
        'Time': [datetime.now().strftime('%H:%M:%S')],
        'Date': [(datetime.today().strftime('%Y-%m-%d'))],
        'WPIndex': [str(wpindex)],
        'FRR': [FRR],  
        'TotalFR': [TotalFR],  
        'Volume': [volume],
        'Eq Time': [eq_t],
        'Buf-Name': [expparams[5]],
        'Buf-FR': [exp_FRs[0]],
        'Buf-FRer': [fr_perc_error[0]],
        'Buf-FRmean': [mean_fr[0]],
        'Buf-FRstd': [stdev_fr[0]],
        'Lp1-Name': [expparams[6]],
        'Lp1-Comp': [expparams[2]],
        'Lp1-FR': [exp_FRs[1]],
        'Lp1-FRer': [fr_perc_error[1]],
        'Lp1-FRmean': [mean_fr[1]],
        'Lp1-FRstd': [stdev_fr[1]],
        'Lp2-Name': [expparams[7]],
        'Lp2-Comp': [expparams[3]],
        'Lp2-FR': [exp_FRs[2]],
        'Lp2-FRer': [fr_perc_error[2]],
        'Lp2-FRmean': [mean_fr[2]],
        'Lp2-FRstd': [stdev_fr[2]],
        'Lp3-Name': [expparams[8]],
        'Lp3-Comp': [expparams[4]],
        'Lp3-FR': [exp_FRs[3]],
        'Lp3-FRer': [fr_perc_error[3]],  
        'Lp3-FRmean': [mean_fr[3]],
        'Lp3-FRstd': [stdev_fr[3]],      
    }

    # Convert data to DataFrame
    df = pd.DataFrame(data)

    # Excel file name
    excel_file = (path + "/flowdata/" + expname + "explog.xlsx")

    try:
        # Load the existing workbook
        book = load_workbook(excel_file)
        
        # Select the active sheet
        sheet = book.active
        
        # Find the next available row
        next_row = sheet.max_row + 1
        
        # Append DataFrame to the existing sheet
        for r in dataframe_to_rows(df, index=False, header=False):
            sheet.append(r)
        
        # Save the workbook
        book.save(excel_file)
        
        print("Data appended to Excel successfully.")
        
    except FileNotFoundError:
        # If the file doesn't exist, create a new Excel file and write the DataFrame
        df.to_excel(excel_file, index=False)
    print("Data written to a new Excel file successfully.")

#-----------------------------  

def initatefolder(path,fpath):
    if not os.path.exists(fpath):
        os.makedirs(fpath)
        print(f"Directory '{path}' created successfully.")
    else:
        print(f"Directory '{path}' already exists.")


def plotupdate(ax1,ax2,flow_data,active_channels,col_list,set_FR,p_range,n,k,single):
    ax1.clear()
    ax2.clear()
    chancolours = ["lightseagreen","darkorange","darkred","blueviolet"]
    for i, chan in enumerate(active_channels):
        ax1.plot(flow_data[0][n:],flow_data[1][i][n:],label="Channel %d" % chan, color=chancolours[i])
        ax2.plot(flow_data[0][n:],flow_data[2][i][n:],label="Channel %d SetP" % chan,color=chancolours[i])
        ax2.plot(flow_data[0][n:],flow_data[3][i][n:],label="Channel %d ActP" % chan,color=chancolours[i], alpha=0.5)
        if single == True:
            ax1.axhline(y=set_FR[i],color=chancolours[i],linewidth=0.5, alpha=0.5)
    if col_list > 0:
        ax1.axvline(x=col_list,color='k',linestyle=':')
        ax2.axvline(x=col_list,color='k',linestyle=':')                    
    ax1.set_ylim(-5, np.amax(set_FR)*1.5) #np.amax(set_FR)*1.5
    ax2.set_ylim(p_range[0],p_range[1])
    ax1.legend()
    ax2.legend()
    ax1.set_ylabel('flow [uL/min]')
    ax2.set_ylabel('pressure [mbar]')
    plt.xlabel('Time (s)')
    plt.pause(0.0001)
    plt.show(block=False)

#------------------
                  
def main_PI(expname,exp_params,autocollect,active_channels,sensorcorr,period,K_p,K_i,exp_FRs,volume,p_range,p_incr,eqb_max,eqb_min,wpdim,wpcurrent,standard_repeats,ser,calibarr):
    
    #Global variables
    flow_data,n,fdataall = [[],[[],[],[],[]],[[],[],[],[]],[[],[],[],[]]],0,[] #time,flow,realp,setp
    pr_control, I = [0,0,0,0],[0,0,0,0]
    init_start_t = time.time()

    #create folder for saving images
    path = './FlowData/' + expname + './Flowplots'
    fpath = './FlowData/' + expname 
    
    initatefolder(path,fpath)

    #Setup live flow figure
    fig = plt.figure()
    ax1 = fig.add_subplot(2,1,1)
    ax2 = fig.add_subplot(2,1,2)
    fr = [0,0,0,0]

    #Iteration through each of the flow rates
    for j, set_FR in enumerate(exp_FRs): 
        
        FRR = set_FR[0]/np.sum(set_FR[1:])
        
        repeats = 0
        col_list = 0

        #exp_name = str(round(exp_params[1],1)) + "-" + expname + str(exp_params[2]) + str(exp_params[3]) + str(exp_params[4] )
        exp_name = str(round(FRR,1)) + "-" + expname + "-" + str(set_FR[0]) + str(set_FR[1]) + str(set_FR[2]) + str(set_FR[3] )
        print(exp_name)
        #caclulate experiment time based off the target volume
        exp_t = (volume/np.sum(set_FR[0:(len(active_channels)-1)]))*60

        #tubevol = ((math.pi)*(tubingdim[0]/2)**2)*(tubingdim[1]*10)

        while repeats < standard_repeats: #Gives the option to repeat a given flowrate condition
            
            print("Beginning experiment: ", set_FR)
            
            fr_perc_error, max_fr_error, error_prev = [[0],[0],[0],[0]],[],[0,0,0,0]
            collection,consistent_fr = False,True

            #Read inital time
            start_t = time.time() 
            last_t = start_t
            active_t = eqb_max           

            while True:
                #-----Main PID Loop-----
                flow_data[0].append((time.time() - init_start_t))

                #Repeat for all active channels
                for i, channel  in enumerate(active_channels):
                    #Flow rate reading
                    fr[i] = pump.get_sensor_data(channel)[0]
                    #Apply Correction
                    fr[i] = (fr[i]*sensorcorr[channel-1][0]) + sensorcorr[channel-1][1]
                    
                    flow_data[1][i].append(fr[i])
                    #Calculate error
                    fr_error = fr[i]-set_FR[i]
                    fr_perc_error[i].append(abs(fr_error/set_FR[i]))

                    if error_prev[i]*fr_error < 0 and len(flow_data[2][i]) > 2:
                       #Perform linear interpolation
                        pr_control[i] = flow_data[2][i][-2] + (((flow_data[2][i][-1]-flow_data[2][i][-2])/(fr_error-error_prev[i]))*(0-error_prev[i]))
                        pr_control[i] = np.clip(pr_control[i], p_range[0], p_range[1])
                        flow_data[2][i].append(pr_control[i])
                    else:
                        #Correction calculations
                        I[i] = I[i] + K_i*fr_error*(period)
                        adjustment =  (fr_error*abs(fr_error))*K_p[i] + I[i]
                        #Set limits
                        adjustment = np.clip(adjustment, p_incr[0], p_incr[1])
                        pr_control[i] = pr_control[i] - adjustment 
                        pr_control[i] = np.clip(pr_control[i], p_range[0], p_range[1])
                        flow_data[2][i].append(pr_control[i])
                    #Set pressure
                    pump.set_pressure(channel,pr_control[i],calibarr)  
                    flow_data[3][i].append(pump.get_pressure_data(channel,calibarr)[0])
                    error_prev[i] = fr_error
                 
                print("1:",fr[0],"2:",fr[1],"3:",fr[2],"4:",fr[3])
                max_fr_error.append(np.max([errors[-1] for errors in fr_perc_error]))

                if (time.time() - start_t) > eqb_min and collection == False:
                    print(np.max(max_fr_error)*set_FR[i])
                    if np.max(max_fr_error)*set_FR[i] < 1:
                        print("FR condition reached \n\n\n")
                        #Move to well position
                        if autocollect == True:
                            expel.servoswitch(ser,1) #to collect mode
                        K_p = K_p/5
                        K_i = K_i/10
                        collection = True
                        col_list = time.time() - init_start_t
                        eq_t = time.time() - start_t
                        active_t = exp_t + time.time() - start_t
                        flowstart = len(flow_data[1][0])
                        fr_perc_error, max_fr_error = [[0],[0],[0],[0]],[]
                    max_fr_error = max_fr_error[1:] #Remove the first value 
    
                elif np.max(max_fr_error) > 0.05 and collection == True:
                    consistent_fr = False
                        
                if (time.time() - start_t) > active_t:
                    if autocollect == True:
                        expel.servoswitch(ser,0) #To waste mode
                    break


                # Wait until desired period time
                sleep_t = period - (time.time() - last_t)
                if sleep_t > 0:
                    time.sleep( sleep_t )
                last_t = time.time() # And update the last time 
                
                #Updating Figure
                plotupdate(ax1,ax2,flow_data,active_channels,col_list,set_FR,p_range,n,-1,True)    

            if collection == False:
                status = "Failed to Eq"
                eq_t = 0
                print("Equilibration Failed - FR percentage error", np.max(max_fr_error))
                repeats += -1
                flowstart = 0
            else:
                if consistent_fr == True: 
                    status = "Complete"
                    print("Collection Successful - FR percentage error", np.max(max_fr_error)) 
                    
                else:
                    status = "Failed"
                    print("Collection Unsuccessful - FR percentage error", np.max(max_fr_error)) 
                    #repeats += -1
            
            repeats+= 1

            n = len(flow_data[0])-1
            ch_fr_error,mean_fr,stdev_fr = saveflowdata(fr_perc_error,flow_data,flowstart,fdataall,path,expname)
            savetoexcel(path,expname,exp_name,status,exp_params,set_FR,wpcurrent,volume,ch_fr_error,mean_fr,stdev_fr,standard_repeats,eq_t)

            plt.savefig(path+"/"+exp_name+"-"+str(j)+"-"+str(repeats)+".png")
            #save wp positions with exp info
            wpprev = [wpcurrent[0],wpcurrent[1]]
            if wpcurrent[1] == wpdim[1]:
                wpcurrent[1] = 1
                wpcurrent[0] = wpcurrent[0] + 1
            else:
                wpcurrent[1] = wpcurrent[1] + 1

            #Move to new coordinate 
            if autocollect == True:
                threading.Thread(target=expel.nextwell,args=(ser, wpprev, wpcurrent)).start()
                print("Current plate position: ",wpcurrent)

            K_p = K_p*5
            K_i = K_i*10

            print(repeats,standard_repeats)
            
    stop(calibarr) 
    if autocollect == True:
        expel.setdirection(ser,"Vert","Away")
        expel.setstep(ser,0,2000)
    plotupdate(ax1,ax2,flow_data,active_channels,col_list,exp_FRs,p_range,0,0,False)    
    plt.savefig(path+"/"+expname+"all"+".png")
    plt.show()

    return()