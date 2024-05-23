import numpy as np
from datetime import datetime
import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import math
from datetime import datetime


def id():
    previd = np.load("C:/Users/bd923/OneDrive - Imperial College London/PhD/GitRepos/auto-LNP/id.npy")
    id = previd + 1
    np.save("C:/Users/bd923/OneDrive - Imperial College London/PhD/GitRepos/auto-LNP/id.npy",id)
    return id

def date():
   now = datetime.now() 
   date = now.strftime("%y%m%d")
   print(date)
   return date

def time():
    now = datetime.now() 
    time = now.strftime("%H:%M:%S")
    print(time)
    return time

def saverecord(ExpTitle,Details,Mode,BufferName,Buffer,BufferNotes,Lp1Name,Lp1,Lp1Notes,Lp2Name,Lp2Notes,Lp2,Lp3Name,Lp3Notes,Lp3,exp_params,exp_FRs,inst_name,active_chans,sensorcorr,volume,standard_repeats,fail_repeats,period,K_p,K_i,p_incr,p_range,max_eq_t,Eqabserror):
    now = datetime.now()
    date = now.strftime("%y%m%d")
    time = now.strftime("%H:%M:%S")
    previd = np.load("C:/Users/bd923/OneDrive - Imperial College London/PhD/GitRepos/auto-LNP/id.npy")
    id = previd + 1
    np.save("C:/Users/bd923/OneDrive - Imperial College London/PhD/GitRepos/auto-LNP/id.npy",id)

    expname = f"{id}-{date}-{ExpTitle}"  

    print("Beginning Experiment:",expname)

    fpath = './FlowData/' + expname 
    
    if not os.path.exists(fpath):
        os.makedirs(fpath)  

    data = {
        'id': [id],
        'Date': [date],
        'Time': [time],
        'expname': [expname],
        'Details': [Details],
        'Mode': [Mode],
        'BufferName': [BufferName],
        'BufferParams': [Buffer],
        'BufferNotes': [BufferNotes],
        'Lp1Name': [Lp1Name],
        'Lp1Params': [Lp1],
        'Lp1Notes': [Lp1Notes],  
        'Lp2Name': [Lp2Name],  
        'Lp2Notes': [Lp2Notes],
        'Lp2Params': [Lp2],
        'Lp3Name': [Lp3Name],
        'Lp3Notes': [Lp3Notes], 
        'Lp3Params': [Lp3],
        'exp_Params': [exp_params],
        'exp_FRs': [exp_FRs],
        'inst_name': [inst_name],
        'active_chans': [active_chans],   
        'sensorcorr': [sensorcorr],
        'volume': [volume],    
        'standard_repeats': [standard_repeats], 
        'fail_repeats': [fail_repeats], 
        'period': [period], 
        'K_p': [K_p], 
        'K_i': [K_i], 
        'p_incr': [p_incr], 
        'p_range': [p_range], 
        'maabs_error': [Eqabserror]
    }

    df = pd.DataFrame(data)
    fname = './FlowData/' + expname + '/param_record-' + expname 
    df_transposed = df.T
    df_transposed.to_csv(fname+".csv")
    print("Parameter record saved to ",fname)

    return(expname)

def saveflowdata(fr_err_clip,flow_data,flowstart,fdataall,path,expname):
    ch_fr_error = [np.max(errors) for errors in fr_err_clip]
    mean_fr = [np.mean(data[flowstart:]) for data in flow_data[1]]
    stdev_fr = [np.std(data[flowstart:]) for data in flow_data[1]]
    fdata = [list(data[flowstart:]) for data in flow_data[1]]
    fdataall.append(fdata)
    DF = pd.DataFrame(fdataall)
    DF.to_csv(path+expname+"-steadyflowonly.csv")
    flowdf = pd.DataFrame(flow_data)
    flowdf.to_csv(path+expname+"-Flowdata.csv")
    return(ch_fr_error,mean_fr,stdev_fr)

def savetoexcel(path,expname,exp_name,status,expparams,exp_FRs,wpindex,volume,fr_perc_error,mean_fr,stdev_fr,repeat,eq_t):
    TotalFR = np.sum(exp_FRs)
    FRR = exp_FRs[0]/np.sum(exp_FRs[1:])
    data = {
        'ExpName': [exp_name],
        'State': [status],
        'RepeatNum': [repeat],
        'Time': [datetime.now().strftime('%H:%M:%S')],
        'Date': [(datetime.today().strftime('%Y-%m-%d'))],
        'WPIndex': [str(wpindex)],
        'FRR': [FRR],  
        'TotalFR': [TotalFR],  
        'Volume': [volume],
        'Eq Time': [eq_t],
        'Buf-Name': [expparams[5]],
        'Buf-FR': [exp_FRs[0]],
        'Buf-FRer': [fr_perc_error[0]],
        'Buf-FRmean': [mean_fr[0]],
        'Buf-FRstd': [stdev_fr[0]],
        'Lp1-Name': [expparams[6]],
        'Lp1-Comp': [expparams[2]],
        'Lp1-FR': [exp_FRs[1]],
        'Lp1-FRer': [fr_perc_error[1]],
        'Lp1-FRmean': [mean_fr[1]],
        'Lp1-FRstd': [stdev_fr[1]],
        'Lp2-Name': [expparams[7]],
        'Lp2-Comp': [expparams[3]],
        'Lp2-FR': [exp_FRs[2]],
        'Lp2-FRer': [fr_perc_error[2]],
        'Lp2-FRmean': [mean_fr[2]],
        'Lp2-FRstd': [stdev_fr[2]],
        'Lp3-Name': [expparams[8]],
        'Lp3-Comp': [expparams[4]],
        'Lp3-FR': [exp_FRs[3]],
        'Lp3-FRer': [fr_perc_error[3]],  
        'Lp3-FRmean': [mean_fr[3]],
        'Lp3-FRstd': [stdev_fr[3]],      
    }

    # Convert data to DataFrame
    df = pd.DataFrame(data)

    # Excel file name
    excel_file = (path + expname + "explog.xlsx")

    try:
        # Load the existing workbook
        book = load_workbook(excel_file)
        
        # Select the active sheet
        sheet = book.active
        
        # Find the next available row
        next_row = sheet.max_row + 1
        
        # Append DataFrame to the existing sheet
        for r in dataframe_to_rows(df, index=False, header=False):
            sheet.append(r)
        
        # Save the workbook
        book.save(excel_file)
        
        print("Data appended to Excel successfully.")
        
    except FileNotFoundError:
        # If the file doesn't exist, create a new Excel file and write the DataFrame
        df.to_excel(excel_file, index=False)
    print("Data written to a new Excel file successfully.")

def initatefolder(path):
    if not os.path.exists(path):
        os.makedirs(path)
        print(f"Directory '{path}' created successfully.")
    else:
        print(f"Directory '{path}' already exists.")
